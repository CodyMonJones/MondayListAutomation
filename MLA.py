import openpyxl
from openpyxl import Workbook

# Used to create a new excel workbook
wb = Workbook()

#Creates each sheet in excel workbook we created above
ws = wb.active
ws1 = wb.create_sheet("MAE")
ws2 = wb.create_sheet("CSE")
ws3 = wb.create_sheet("EE")
ws4 = wb.create_sheet("IE")
ws5 = wb.create_sheet("BE")
ws6 = wb.create_sheet("CE")

#Declaring the Title for each worksheet
ws.title = "IGNORE"

"""

Anytime we modify the workbook object, sheets or cells
we need to call save() otherwise the changes will not be made
in our excel file

"""
wb.save("/Users/codyjones/Documents/demo.xlsx")
