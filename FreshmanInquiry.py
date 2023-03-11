import openpyxl
from openpyxl import Workbook

# Used to create a new excel workbook
wb = Workbook()

#Creates each sheet in excel workbook we created above
ws = wb.active #This sheet is a default sheet created when making a new workbook
ws.title = "IGNORE"

#Titles for our subsequent worksheets
ws1 = wb.create_sheet("MAE")
ws2 = wb.create_sheet("CSE")
ws3 = wb.create_sheet("EE")
ws4 = wb.create_sheet("IE")
ws5 = wb.create_sheet("BE")
ws6 = wb.create_sheet("CE")

#Title Bar for each Column
for sheet in wb:
    sheet["A1"] =  "Ambassador Assigned"
    sheet["B1"] =  "Emailed On"
    sheet["C1"] =  "Emailed By"
    sheet["D1"] =  "Comments"
    sheet["E1"] =  "First Name"
    sheet["F1"] =  "Preferred Name"
    sheet["G1"] =  "Last Name"
    sheet["H1"] =  "Email"
    sheet["I1"] =  "Start Term"
    sheet["J1"] =  "Major 1"
    sheet["K1"] =  "Major 2"
    sheet["L1"] =  "Major 3"

"""

Anytime we modify the workbook object, sheets or cells
we need to call save() otherwise the changes will not be made
in our excel file

"""
wb.save("/Users/codyjones/Documents/demo.xlsx")
